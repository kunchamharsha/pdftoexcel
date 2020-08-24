import os 
from bs4 import BeautifulSoup 
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Alignment
import argparse

from openpyxl.styles import Font

import uuid

def generate_word_map(source):
    file_name=os.path.basename(source)
    map_path='maps/'+os.path.splitext(file_name)[0]+'.html'
    command="pdftotext -f 1 -l 1 -r 300 -nodiag -layout -bbox-layout {0} {1}".format(source,map_path)
    try:
        os.system(command)
    except Exception:
        raise("Error generating config file")
    return map_path

def parse_map_to_table(map_path):
    parsed_data=[]
    x=[]
    y=[]
    text={}
    min_cood=[]
    max_cood=[]
    textmap={}
    minmap={}
    maxminmap={}
    parse_map=defaultdict(str)
    corelation_map=defaultdict(dict)
    with open(map_path, 'r') as f:
        html_data = f.read()
        soup = BeautifulSoup(html_data, 'html.parser')
        flows=soup.find_all('flow')
        for flow in flows:
            blocks=flow.find_all('block')
            count=0
            for block in blocks:
                for line in block.find_all('line'):
                    count+=1
                    words=[word.string for word in line.find_all('word')]
                    text=' '.join(words)
                    ymax=0
                    xmax=0
                    ymax=float(line.get('ymax'))
                    xmax=float(line.get('xmax'))
                    xmin=float(line.get('xmin'))
                    ymin=float(line.get('ymin'))
                    maxminmap[(xmax,ymax)]=[xmin,ymin]
                    min_cood.append((float(xmin),float(ymin)))
                    max_cood.append((float(xmax),float(ymax)))
                    y.append(ymax)
                    x.append(xmax)
                    parse_map[str(xmax)+","+str(ymax)]=str(text)
    return parse_map,list(set(x)),list(set(y)),min_cood,max_cood,maxminmap

def create_excel_sheet(word_map,xcood,ycood,min_cood,max_cood,source,maxminmap):
    
    file_name=os.path.basename(source)
    sheet_path='final/'+os.path.splitext(file_name)[0]+'.xlsx'
    workbook = Workbook()
    sheet = workbook.worksheets[0]
    ycood.sort()
    xcood.sort()
    row=0
    count=0
    rowid=1
    colid=1
    colrelindex={}
    maxcount=0
    headersrow=1
    for y in ycood:
        for x in xcood:
            try:
                if (x,y) in max_cood:
                    xm,ym=maxminmap[(x,y)]
                    position=colrelindex.get(xm-xm%100) or colrelindex.get((xm/100)*100)
                    if position!=None:
                        sheet.cell(row=rowid, column=position).value=word_map[str(x)+","+str(y)]
                    else:
                        count+=1
                        sheet.cell(row=rowid, column=colid).value=word_map[str(x)+","+str(y)]
                        if xm-xm%100<50:
                            colrelindex[xm-x%100]=colid
                        else:
                            colrelindex[int(xm/100)*100]=colid
                    colid+=1
                    count+=1
            except ValueError:
                sheet.cell(row=rowid, column=colid).value=""
        if count>maxcount:
            maxcount=count
            headersrow=rowid
        rowid+=1
        colid=1
        count=0
    font = Font(color="FF0000")
    
    for indexrow,row in enumerate(sheet.iter_rows(),1):
        for cell in row:
            if indexrow==headersrow:
                cell.font = font
                sheet.freeze_panes=cell
            cell.alignment = Alignment(wrapText=True)
    try:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrapText=True)
        workbook.save(filename=sheet_path)
    except Exception as e:
        print(e)
    return sheet_path


if __name__ == "__main__":
    source="test.pdf"
    parser = argparse.ArgumentParser()
    parser.add_argument('--source', help='add source of pdf file')
    args = parser.parse_args()
    source=args.source
    if source==None:
        print("Mention source using --source flag")
    else:
        map_path=generate_word_map(source)
        word_map,xcood,ycood,min_cood,max_cood,maxminmap=parse_map_to_table(map_path)
        excel_sheet=create_excel_sheet(word_map,xcood,ycood,min_cood,max_cood,source,maxminmap)

    